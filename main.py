import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import schedule
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime

# 配置日志记录
logging.basicConfig(filename='comments.log', level=logging.INFO, format='%(asctime)s - %(message)s')

live_dy_url = 'https://live.douyin.com/627222682278'

options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
service = Service()

browse = webdriver.Chrome(service=service, options=options)

# 用于存储已处理的评论内容
seen_comments = set()

# Excel 文件路径
excel_file = 'comments.xlsx'

# 如果文件不存在，创建新的工作簿
try:
    wb = load_workbook(excel_file)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["时间", "评论"])  # 添加表头
    wb.save(excel_file)

browse.get(live_dy_url)
time.sleep(10)  # 等待页面加载
def fetch_comments():
    try:

        comments = browse.find_elements(By.XPATH, ".//*[@class='webcast-chatroom___content-with-emoji-text']")
        for comment in comments:
            comment_text = comment.text
            if comment_text not in seen_comments:  # 检查评论是否已经处理过
                seen_comments.add(comment_text)  # 将新评论添加到集合中
                print(comment_text)
                logging.info(comment_text)  # 记录评论到日志文件
                # 将评论和时间戳写入 Excel
                ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), comment_text])
                wb.save(excel_file)
    except Exception as e:
        print(f"Error fetching comments: {e}")
        logging.error(f"Error fetching comments: {e}")

# 每隔10秒抓取一次评论
schedule.every(10).seconds.do(fetch_comments)

while True:
    schedule.run_pending()
    time.sleep(1)